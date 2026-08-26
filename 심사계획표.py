import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "audit.db")
EXCEL_SRC_PATH = os.path.join(BASE_DIR, "정비본부 품질심사 실시이력(2026.XX.XX).xlsm")
OUTPUT_EXCEL_PATH = os.path.join(BASE_DIR, "품질심사_스케줄계획_2026_2030.xlsx")

def get_db_connection():
    return sqlite3.connect(DB_PATH)

def initialize_database():
    """DB와 테이블이 없으면 초기 마이그레이션을 수행합니다."""
    if not os.path.exists(DB_PATH):
        print("[시스템] 데이터베이스가 존재하지 않아 초기 마이그레이션을 실행합니다...")
        import sys
        sys.path.append(os.path.dirname(os.path.abspath(__file__)))
        try:
            # db_migration.py 실행
            migration_script = r"C:\Users\twayair\.gemini\antigravity-ide\brain\c1badfe7-8c6e-4afd-8ac0-ef94f9e18f7e\scratch\db_migration.py"
            if not os.path.exists(migration_script):
                migration_script = "db_migration.py"
            
            # 직접 마이그레이션 함수를 호출하거나 스크립트 실행
            os.system(f"python \"{migration_script}\"")
        except Exception as e:
            print(f"[오류] 초기 마이그레이션 실패: {e}")

def get_all_targets():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, category, operation_status, station_name, base_interval FROM audit_targets")
    rows = cursor.fetchall()
    conn.close()
    return [{"id": r[0], "category": r[1], "operation_status": r[2], "station_name": r[3], "base_interval": r[4]} for r in rows]

def add_schedule_event(target_id, year, event_type, detail):
    """신규 지점/기종 추가 이벤트를 등록합니다."""
    conn = get_db_connection()
    cursor = conn.cursor()
    # 중복 삽입 방지
    cursor.execute("""
        SELECT id FROM schedule_events 
        WHERE target_id = ? AND event_year = ? AND event_type = ?
    """, (target_id, year, event_type))
    if cursor.fetchone():
        cursor.execute("""
            UPDATE schedule_events SET detail = ? 
            WHERE target_id = ? AND event_year = ? AND event_type = ?
        """, (detail, target_id, year, event_type))
        print(f"[DB] 기존 이벤트 업데이트 완료 ({year}년, Target ID: {target_id})")
    else:
        cursor.execute("""
            INSERT INTO schedule_events (target_id, event_year, event_type, detail)
            VALUES (?, ?, ?, ?)
        """, (target_id, year, event_type, detail))
        print(f"[DB] 신규 이벤트 등록 완료 ({year}년, Target ID: {target_id})")
    conn.commit()
    conn.close()

def generate_schedule_for_target(target_id, start_year=2026, end_year=2030):
    """
    특정 대상에 대해 2026~2030년 스케줄링 알고리즘을 수행합니다.
    - 2년마다 직접/간접 순차 교대 실시.
    - 지점/기종 추가 이벤트가 지정된 연도에는 '간접' 강제 편성 및 주기 리셋.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. 2025년 이전의 가장 최근 심사 완료 기록 가져오기
    cursor.execute("""
        SELECT year, audit_type FROM audit_history 
        WHERE target_id = ? AND year < ? AND status = '완료'
        ORDER BY year DESC LIMIT 1
    """, (target_id, start_year))
    last_audit = cursor.fetchone()
    
    # 2. 해당 기간(start_year ~ end_year)의 스케줄 이벤트 가져오기
    cursor.execute("""
        SELECT event_year, detail FROM schedule_events
        WHERE target_id = ? AND event_year BETWEEN ? AND ?
    """, (target_id, start_year, end_year))
    events = {r[0]: r[1] for r in cursor.fetchall()}
    conn.close()
    
    schedule = {}
    
    # 기준점 초기화
    if last_audit:
        ref_year, ref_type = last_audit[0], last_audit[1]
    else:
        ref_year, ref_type = None, None
        
    for year in range(start_year, end_year + 1):
        # 1) 기종/지점 추가 등 변동 이벤트가 있는 연도
        if year in events:
            audit_type = "간접"
            detail = events[year]
            schedule[year] = {
                "audit_type": audit_type,
                "is_event": True,
                "detail": detail,
                "symbol": "◎"
            }
            # 기준점 리셋 (간접 심사로 갱신)
            ref_year = year
            ref_type = "간접"
            
        # 2) 이벤트는 없지만 정기 주기가 도래한 연도 (기준년도로부터 2년 뒤)
        elif ref_year is not None:
            if year - ref_year == 2:
                audit_type = "간접" if ref_type == "직접" else "직접"
                symbol = "◎" if audit_type == "간접" else "●"
                schedule[year] = {
                    "audit_type": audit_type,
                    "is_event": False,
                    "detail": "정기 심사",
                    "symbol": symbol
                }
                # 기준점 갱신 (이번 심사 편성 연도로 기준 이동)
                ref_year = year
                ref_type = audit_type
            else:
                schedule[year] = None
        else:
            # 과거 이력이 없는 경우, start_year에 최초 직접심사 시작
            if year == start_year:
                audit_type = "직접"
                schedule[year] = {
                    "audit_type": audit_type,
                    "is_event": False,
                    "detail": "최초 직접 심사",
                    "symbol": "●"
                }
                ref_year = year
                ref_type = audit_type
            else:
                schedule[year] = None
                
    return schedule

def run_scheduling_simulation(start_year=2026, end_year=2030):
    """전체 대상에 대한 스케줄링 시뮬레이션을 수행하고 결과를 반환합니다."""
    targets = get_all_targets()
    results = []
    
    for t in targets:
        sched = generate_schedule_for_target(t["id"], start_year, end_year)
        row_data = {
            "id": t["id"],
            "category": t["category"],
            "operation_status": t["operation_status"],
            "station_name": t["station_name"]
        }
        for year in range(start_year, end_year + 1):
            val = sched.get(year)
            if val:
                row_data[f"{year}_symbol"] = val["symbol"]
                row_data[f"{year}_detail"] = val["detail"]
            else:
                row_data[f"{year}_symbol"] = ""
                row_data[f"{year}_detail"] = ""
        results.append(row_data)
        
    return pd.DataFrame(results)

def export_to_excel(df, filename=OUTPUT_EXCEL_PATH):
    """편성된 스케줄 결과를 미려하게 서식화된 엑셀 파일로 출력합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "품질심사 연간 계획"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Malgun Gothic", size=16, bold=True, color="1F497D")
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Malgun Gothic", size=9)
    symbol_font = Font(name="Malgun Gothic", size=11, bold=True)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    direct_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # 연한 블루
    indirect_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid") # 연한 주황
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # 1. Title
    ws["B2"] = "정비본부 품질심사 연간 일정 계획 (2026 - 2030)"
    ws["B2"].font = title_font
    ws.row_dimensions[2].height = 30
    
    # 2. Table Headers
    # Row 4: Category, Status, Station, Years merged (2 cols per year)
    headers_row4 = ["유형", "운항여부", "심사대상 (Station)"]
    years = [2026, 2027, 2028, 2029, 2030]
    
    for h in headers_row4:
        col_num = len(ws[4]) + 1
        cell = ws.cell(row=4, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        # Merge vertical cells for row 4 and row 5
        ws.merge_cells(start_row=4, start_column=col_num, end_row=5, end_column=col_num)
        
    for y in years:
        start_col = len(ws[4]) + 1
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col+1)
        cell = ws.cell(row=4, column=start_col, value=f"{y}년 계획")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Subheaders in Row 5
        c1 = ws.cell(row=5, column=start_col, value="구분")
        c2 = ws.cell(row=5, column=start_col+1, value="세부 일정 / 비고")
        for c in [c1, c2]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    # 3. Populate Data
    start_row = 6
    for idx, row in df.iterrows():
        # Insert base info
        ws.cell(row=start_row, column=1, value=row["category"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row, column=2, value=row["operation_status"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row, column=3, value=row["station_name"]).alignment = Alignment(horizontal="left", vertical="center")
        
        col_idx = 4
        for y in years:
            sym = row[f"{y}_symbol"]
            det = row[f"{y}_detail"]
            
            c_sym = ws.cell(row=start_row, column=col_idx, value=sym)
            c_det = ws.cell(row=start_row, column=col_idx+1, value=det)
            
            c_sym.alignment = Alignment(horizontal="center", vertical="center")
            c_det.alignment = Alignment(horizontal="left", vertical="center")
            
            # Styling based on type
            if sym == "●":
                c_sym.fill = direct_fill
                c_sym.font = symbol_font
            elif sym == "◎":
                c_sym.fill = indirect_fill
                c_sym.font = symbol_font
                
            col_idx += 2
            
        # Apply border & fonts to all cells in the row
        for col in range(1, col_idx):
            cell = ws.cell(row=start_row, column=col)
            cell.border = thin_border
            if cell.font.name != "Malgun Gothic":
                cell.font = data_font
                
        ws.row_dimensions[start_row].height = 22
        start_row += 1

    # Auto-adjust column widths
    for col in range(1, col_idx):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(4, start_row):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Extra width for Station column
    ws.column_dimensions["C"].width = 30
    
    wb.save(filename)
    print(f"[엑셀] 편성표가 다음 경로에 성공적으로 저장되었습니다: {filename}")

def print_eastec_scenario_preview():
    """사용자가 요청한 EASTEC B737-8 MAX 기종 추가 시나리오 시뮬레이션 결과를 출력합니다."""
    print("\n" + "="*80)
    print(" [시뮬레이션 검증] EASTEC 위탁정비 기종 추가 시나리오 (B737-8 MAX, 2027년 추가)")
    print("="*80)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    # EASTEC 타겟 찾기
    cursor.execute("SELECT id, station_name, category FROM audit_targets WHERE station_name LIKE '%EASTEC%'")
    eastec_targets = cursor.fetchall()
    conn.close()
    
    if not eastec_targets:
        print("[!] DB에 EASTEC 관련 타겟이 검색되지 않습니다. 먼저 DB 마이그레이션을 확인하세요.")
        return
        
    print(f"-> 매핑된 EASTEC 대상 목록:")
    for t_id, name, cat in eastec_targets:
        print(f"   [ID: {t_id}] {cat} - {name}")
        
    # EASTEC (베이징) 또는 예제용 EASTEC 상하이 임시 등록
    # 시나리오 재현을 위해 "EASTEC (상하이)" 지점을 타겟으로 새로 추가하고 시뮬레이션
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. 2026년 직접 수행 이력이 이미 있는 'EASTEC (상하이)' 타겟 등록 (데모용)
    cursor.execute("SELECT id FROM audit_targets WHERE station_name = 'EASTEC (상하이)'")
    existing = cursor.fetchone()
    if not existing:
        cursor.execute("""
            INSERT INTO audit_targets (category, operation_status, station_name, base_interval)
            VALUES ('완전위탁업체', '운항', 'EASTEC (상하이)', 2)
        """)
        demo_target_id = cursor.lastrowid
        # 2026년 직접 심사 완료 이력 추가
        cursor.execute("""
            INSERT INTO audit_history (target_id, year, audit_type, status, remarks)
            VALUES (?, 2026, '직접', '완료', '2026년 정기 직접 심사 완료')
        """, (demo_target_id,))
        conn.commit()
        print(f"-> 데모용 'EASTEC (상하이)' 신규 등록 완료 (ID: {demo_target_id}, 2026년 직접심사 완료 상태)")
    else:
        demo_target_id = existing[0]
        # 혹시 기존 2026년 이력 확인
        cursor.execute("SELECT id FROM audit_history WHERE target_id = ? AND year = 2026", (demo_target_id,))
        if not cursor.fetchone():
            cursor.execute("""
                INSERT INTO audit_history (target_id, year, audit_type, status, remarks)
                VALUES (?, 2026, '직접', '완료', '2026년 정기 직접 심사 완료')
            """, (demo_target_id,))
            conn.commit()
    conn.close()

    # 이벤트 적용 전 스케줄 출력
    print("\n1. [이벤트 적용 전] 2026~2030년 예상 스케줄 (기본 2년 주기):")
    # 기존 이벤트가 있다면 제거
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM schedule_events WHERE target_id = ? AND event_year = 2027", (demo_target_id,))
    conn.commit()
    conn.close()
    
    sched_before = generate_schedule_for_target(demo_target_id, 2026, 2030)
    for y in range(2026, 2031):
        info = sched_before.get(y)
        symbol = info["symbol"] if info else "-"
        detail = info["detail"] if info else ""
        print(f"   - {y}년: [{symbol}] {detail}")

    # 2027년 'B737-8 MAX 기종 추가 위탁정비' 이벤트 등록
    print("\n2. [이벤트 등록] 2027년 EASTEC (상하이) B737-8 MAX 기종 추가 이벤트 삽입")
    add_schedule_event(demo_target_id, 2027, "FLEET_ADDED", "B737-8 MAX 기종 추가")
    
    # 이벤트 적용 후 스케줄 출력
    print("\n3. [이벤트 적용 후] 2026~2030년 자동 재편성 스케줄:")
    sched_after = generate_schedule_for_target(demo_target_id, 2026, 2030)
    for y in range(2026, 2031):
        info = sched_after.get(y)
        symbol = info["symbol"] if info else "-"
        detail = info["detail"] if info else ""
        print(f"   - {y}년: [{symbol}] {detail}")
        
    print("="*80)

def main_menu():
    initialize_database()
    
    while True:
        print("\n" + "="*50)
        print(" 품질심사 스케줄러 시스템 CLI")
        print("="*50)
        print(" 1. 전체 심사 대상 목록 조회")
        print(" 2. 2026~2030년 전체 자동 스케줄링 시뮬레이션 및 엑셀 출력")
        print(" 3. 신규 지점/기종 추가 이벤트 등록")
        print(" 4. EASTEC 상하이 기종 추가 시나리오 검증 및 시뮬레이션")
        print(" 5. 종료")
        print("="*50)
        
        choice = input(" 메뉴 선택 (1-5): ").strip()
        
        if choice == "1":
            targets = get_all_targets()
            print(f"\n[목록] 총 {len(targets)}개의 심사 대상이 등록되어 있습니다.")
            for t in targets[:20]:
                print(f" - ID: {t['id']} | [{t['category']}] {t['station_name']} ({t['operation_status']})")
            if len(targets) > 20:
                print(f" ... 외 {len(targets)-20}개 대상이 더 존재합니다.")
                
        elif choice == "2":
            print("\n[스케줄링] 2026~2030년 자동 편성 시뮬레이션을 진행합니다...")
            df = run_scheduling_simulation(2026, 2030)
            export_to_excel(df)
            print("[알림] 전체 시뮬레이션 결과가 엑셀로 저장되었습니다.")
            
        elif choice == "3":
            targets = get_all_targets()
            target_name = input(" 이벤트를 추가할 대상(Station/업체명) 검색어 입력: ").strip()
            matched = [t for t in targets if target_name.lower() in t["station_name"].lower()]
            
            if not matched:
                print("[!] 검색된 대상이 없습니다.")
                continue
                
            print("\n[검색 결과]")
            for idx, t in enumerate(matched):
                print(f"  {idx + 1}. ID: {t['id']} | [{t['category']}] {t['station_name']}")
                
            sel_idx = input(" 대상을 선택하세요 (번호 입력): ").strip()
            try:
                sel_target = matched[int(sel_idx) - 1]
                year = int(input(" 이벤트 발생 연도 (예: 2027): ").strip())
                detail = input(" 이벤트 내용 (예: B737-8 MAX 기종 추가): ").strip()
                add_schedule_event(sel_target["id"], year, "FLEET_ADDED", detail)
            except (ValueError, IndexError):
                print("[!] 올바른 번호와 연도를 입력하세요.")
                
        elif choice == "4":
            print_eastec_scenario_preview()
            
        elif choice == "5":
            print("[시스템] 프로그램을 종료합니다.")
            break
        else:
            print("[!] 올바른 메뉴 번호를 입력하세요.")

if __name__ == "__main__":
    # 스크립트 실행 시 CLI 메뉴를 실행합니다.
    # 사용자가 바로 결과를 보고 엑셀로 뽑을 수 있도록 디폴트로 시뮬레이션 및 엑셀 출력을 수행해 줍니다.
    print("[시스템] 품질심사 자동 스케줄링 프로그램을 로드합니다.")
    initialize_database()
    
    # 엑셀 내보내기 자동 수행
    print("[시스템] 2026~2030년 품질심사 기본 스케줄 생성을 실행합니다...")
    df = run_scheduling_simulation(2026, 2030)
    export_to_excel(df)
    
    # 데모 검증 실행
    print_eastec_scenario_preview()
    
    # CLI 메뉴 시작
    main_menu()
