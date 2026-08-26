import openpyxl
import sqlite3
import datetime
import re
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "(동재) 정비본부 품질심사 실시이력(2026.XX.XX).xlsm")
DB_PATH = os.path.join(BASE_DIR, "audit.db")

# 1년 주기 직접심사 대상 키워드 및 공항 목록
ANNUAL_DIRECT_TARGETS = [
    "CJU (제주)", "TAE (대구)", "KWJ (광주)", "PUS (부산)", "PUS (김해)", "CJJ (청주)",
    "대동금속", "고암에이스항공", "항공기기술", "엔진기술", "정비기획", "정비자재", "정비통제", "운항정비1,2", "부품수리"
]

def parse_date_val(val, default_year=None):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    
    val_str = str(val).strip()
    if not val_str or val_str in ["-", "N/A", "N.A", "직접심사", "서류예정", "차기간접"]:
        return None
        
    # Pattern like "2025.10.15~16" or "2025.02.24~28" or "2024-05-21~24"
    m = re.search(r'(\d{4})[.-](\d{1,2})[.-](\d{1,2})', val_str)
    if m:
        y, m_val, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{m_val:02d}-{d:02d}"
        
    # Pattern like "2025.12" or "2023-06"
    m_ym = re.search(r'(\d{4})[.-](\d{1,2})', val_str)
    if m_ym:
        y, m_val = int(m_ym.group(1)), int(m_ym.group(2))
        return f"{y:04d}-{m_val:02d}-15"
        
    # Pattern like "2019-05"
    m_y = re.search(r'(\d{4})', val_str)
    if m_y:
        y = int(m_y.group(1))
        return f"{y:04d}-06-01"
        
    if default_year:
        return f"{default_year}-06-01"
    return None

def import_data():
    print(f"Loading workbook (formula mode): {EXCEL_PATH}")
    wb_formula = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    sheet_f = wb_formula['수행실적']
    
    print(f"Loading workbook (data mode): {EXCEL_PATH}")
    wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet_d = wb_data['수행실적']
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 1. Reset tables with enhanced schema
    cursor.execute("DROP TABLE IF EXISTS schedule_events")
    cursor.execute("DROP TABLE IF EXISTS audit_history")
    cursor.execute("DROP TABLE IF EXISTS audit_targets")
    
    cursor.execute("""
    CREATE TABLE audit_targets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT NOT NULL,
        operation_status TEXT NOT NULL,
        station_name TEXT NOT NULL,
        base_interval INTEGER DEFAULT 2
    )
    """)
    
    cursor.execute("""
    CREATE TABLE audit_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        target_id INTEGER NOT NULL,
        year INTEGER NOT NULL,
        audit_type TEXT NOT NULL,
        status TEXT NOT NULL,
        is_event INTEGER DEFAULT 0,
        event_detail TEXT,
        remarks TEXT,
        scheduled_date TEXT,
        auditor TEXT,
        FOREIGN KEY (target_id) REFERENCES audit_targets(id)
    )
    """)
    
    cursor.execute("""
    CREATE TABLE schedule_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        target_id INTEGER NOT NULL,
        event_year INTEGER NOT NULL,
        event_type TEXT NOT NULL,
        detail TEXT NOT NULL,
        FOREIGN KEY (target_id) REFERENCES audit_targets(id)
    )
    """)
    
    # Year columns mapping from row 5
    year_cols = []
    for c in range(4, 25, 2):
        yr_val = sheet_f.cell(5, c).value
        if yr_val:
            year_cols.append((c, yr_val))
            
    print("Detected Year Columns:", year_cols)
    
    target_count = 0
    history_count = 0
    
    for r in range(6, sheet_f.max_row + 1):
        cat = sheet_f.cell(r, 1).value
        status = sheet_f.cell(r, 2).value
        station = sheet_f.cell(r, 3).value
        
        if not (cat or status or station):
            continue
            
        cat = str(cat).strip() if cat else "해외운항지점"
        status = str(status).strip() if status else "운항"
        station = str(station).strip() if station else ""
        
        # Check base_interval (1 year for domestic station & internal teams)
        is_annual = False
        for kw in ANNUAL_DIRECT_TARGETS:
            if kw in station or station in kw:
                is_annual = True
                break
        base_interval = 1 if is_annual else 2
        
        cursor.execute("""
            INSERT INTO audit_targets (category, operation_status, station_name, base_interval)
            VALUES (?, ?, ?, ?)
        """, (cat, status, station, base_interval))
        target_id = cursor.lastrowid
        target_count += 1
        
        # Parse history and manual plans
        for col_idx, yr_label in year_cols:
            formula_val = sheet_f.cell(r, col_idx).value
            data_val = sheet_d.cell(r, col_idx).value
            detail_val = sheet_d.cell(r, col_idx + 1).value
            
            # Parse Year
            try:
                yr_int = int(yr_label)
            except ValueError:
                if yr_label == "5년전":
                    date_match = parse_date_val(data_val) or parse_date_val(detail_val)
                    if date_match:
                        yr_int = int(date_match.split("-")[0])
                    else:
                        continue
                else:
                    continue
                    
            # Check if this cell is a formula or a manual entry
            is_formula = str(formula_val or "").startswith("=")
            
            # For 2026 and later, only save if it's a manual entry (not a formula) or has custom detail
            if yr_int >= 2026:
                if is_formula and not detail_val:
                    # Let the dynamic scheduler handle formulas!
                    continue
                    
            sym = str(data_val).strip() if data_val is not None else ""
            detail = str(detail_val).strip() if detail_val is not None else ""
            
            # Determine audit type
            audit_type = None
            if "●" in sym:
                audit_type = "직접"
            elif "◎" in sym:
                audit_type = "간접"
            elif "직접" in detail:
                audit_type = "직접"
            elif "간접" in detail:
                audit_type = "간접"
                
            if not audit_type and not sym and not detail:
                continue
                
            if not audit_type:
                if yr_int <= 2025 and (sym or detail):
                    audit_type = "직접"
                else:
                    continue
                    
            status_str = "완료" if yr_int <= 2025 else "계획"
            scheduled_date = parse_date_val(detail_val, default_year=yr_int) or parse_date_val(data_val, default_year=yr_int) or f"{yr_int}-06-01"
            remarks = detail if detail and detail not in ["직접심사", "서류예정", "차기간접"] else ""
            if "서류예정" in detail:
                remarks = "서류예정"
                
            cursor.execute("""
                INSERT INTO audit_history (target_id, year, audit_type, status, scheduled_date, auditor, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (target_id, yr_int, audit_type, status_str, scheduled_date, "", remarks))
            history_count += 1
            
    conn.commit()
    conn.close()
    
    print(f"Successfully imported {target_count} targets and {history_count} history/manual plan records into audit.db!")

if __name__ == "__main__":
    import_data()
